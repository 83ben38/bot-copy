from openpyxl import load_workbook
import rag
import threading
import markdown
from config.scoringconfig import SCORING_CRITERIA

wb = load_workbook('redteaming.xlsx')


##set workbook to the input file



ws = wb.active



def process_question(question_ID):
    question_ID += 1

    QuestionCellID = 'C' + str(question_ID)
    question = ws[QuestionCellID].value

    print("Pulling question from cell: " + QuestionCellID)

    ResponseCellID = 'E' + str(question_ID)



    vector = rag.get_vector_reponse(question)
    
    # Take the question, and process it with 
    response_message = rag.get_chatgpt_response(question, "", rag.get_main_bot_constitution(), vector)
    if (len(response_message) < 5):
        response_message = "I'm sorry, but I can't help with that. I can provide guidance to those who need help with harassment and information about sexual exploitation or similar topics."

    # Save the response
    ws[ResponseCellID] = str(response_message)

    # Score the response
    scores_dict = rag.score_material(response_message, question,vector)
    
    z = 0
    for criterion, description in SCORING_CRITERIA.items():
        ScoreCellID = chr(ord('F') + (z*2))+str(question_ID)
        ReasoningCellID = chr(ord('G') + (z*2))+str(question_ID)
        z+=1
        ws[ScoreCellID] = scores_dict[criterion]['score']
        ws[ReasoningCellID] = scores_dict[criterion]['reasoning']
    # Save the scores
    ws['R' + str(question_ID)] = vector

    # Save in history file
    response_message_html = markdown.markdown(response_message)
    response_message_without_enter = response_message_html.replace("\n",'`')
    fileName = "Question " + str(question_ID-1)
    newFile = open("./history/"+fileName.replace("\n","")+".txt",'w')
    newFile.write(fileName+"\n")
    newFile.write(question+"\n"+response_message_without_enter+"\n"+str(scores_dict)+"\n")
    newFile.close()

    return question_ID

import threading
import queue

start_ID = 1
end_ID = 263
ids = [14,15,17,19,20,23,24,25,26,43,47,49,60,66,68,70,71,75,77,79,80,81,82,83,84,86,88,89,91,92,93,94,96,99,100,102,104,106,107,108,109,110,111,112,113,116,117,118,119,120,121,134,138,144,145,151,163,165,166,167,169,172,173,174,175,176,177,179,180,182,183,184,188,190,195,196,197,198,200,201,203,204,214,215,217,219,222,247,252,254,258,259,260]
def worker(q):
    while not q.empty():
        question_ID = q.get()
        try:
            process_question(question_ID)
        finally:
            q.task_done()
    wb.save("output.xlsx")

# Create a queue and populate it with question IDs
q = queue.Queue()
for question_ID in range(start_ID,end_ID):
    q.put(question_ID)

# Create and start threads
num_threads = 6
threads = []
for _ in range(num_threads):
    thread = threading.Thread(target=worker, args=(q,))
    threads.append(thread)
    thread.start()


for thread in threads:
    thread.join()
    print("Thread completed.\n")

# Save the workbook after all threads have finished processing

